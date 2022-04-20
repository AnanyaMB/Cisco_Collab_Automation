import openpyxl
import paramiko
import time
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook


def cert_identity_details(cert_identity, remote_conn):
    print("------Starting to fetch details of : ", cert_identity)
    expiringIn = None
    isExpired = None
    serialNumber = None
    remote_conn.send('show cert own ' + cert_identity + '\n')
    time.sleep(10)
    identityCert_detail = remote_conn.recv(5000)
    identityCert_detail = identityCert_detail.decode()
    identityCert_detail = identityCert_detail.split("\n")
    remote_conn.send('q')
    time.sleep(5)
    valid_from = None
    valid_to = None
    for date in identityCert_detail:
        date = date.strip()
        if date.startswith('Validity'):
            valid_from = date[15:]
        if date.startswith('To:'):
            valid_to = date.strip("To:")
            valid_to = valid_to.strip()
            date_string = valid_to.replace(valid_to[-8:-5], '').rstrip()
            date_string = date_string.strip()
            valid_to = date_string
            datetime_object = datetime.strptime(
                date_string, '%a %b %d %H:%M:%S %Y').strftime("%Y-%m-%d %H:%M:%S")
            datetime_object = datetime.fromisoformat(datetime_object)
            currentDateTime = datetime.utcnow()
            expiringIn = str(datetime_object - currentDateTime)
            if expiringIn > '0':
                isExpired = False
            else:
                isExpired = True
    for sno in identityCert_detail:
        sno = sno.strip()
        if sno.startswith('Serial Number'):
            serialNumber = sno.strip('Serial Number:')
            serialNumber = serialNumber.strip()
    serviceName = cert_identity.split('/')[0]
    identity_cert_dict = {
        "name": cert_identity,
        "trackingId": serviceName + "_id",
        "serialNumber" : serialNumber,
        "service": serviceName,
        "type": "IdentityCertificate",
        "validFrom": valid_from,
        "validTo": valid_to,
        "expiringIn": expiringIn,
        "hasExpired": isExpired,
    }
    return identity_cert_dict


def fetch_identity_certs(remote_conn, identity_cert_list, identity_cert_dict):
    remote_conn.send('show cert list own\n')
    time.sleep(10)
    identityCert = remote_conn.recv(50000)
    identityCert = identityCert.decode()
    identityCert = identityCert.split("\r\n")
    for eachIdentityCert in identityCert:
        if ".pem" in eachIdentityCert:
            identity_cert_name = eachIdentityCert.split(":")[0]
            identity_cert_details = cert_identity_details(
                identity_cert_name, remote_conn)
            identity_cert_list.append(identity_cert_details)
    #print("Identity List  : ", identity_cert_list)
    return identity_cert_list


def cert_trust_details(cert_trust, remote_conn):
    print("------Starting to fetch details of : ", cert_trust)
    expiringIn = None
    isExpired = None
    valid_from = None
    valid_to = None
    serialNumber = None
    remote_conn.send('show cert trust ' + cert_trust + '\n')
    time.sleep(10)
    trustCert_detail = remote_conn.recv(5000)
    trustCert_detail = trustCert_detail.decode()
    trustCert_detail = trustCert_detail.split("\n")
    remote_conn.send('q')
    time.sleep(5)
    for date in trustCert_detail:
        date = date.strip()
        if date.startswith('Validity'):
            valid_from = date[15:]
        if date.startswith('To:'):
            valid_to = date.strip("To:")
            valid_to = valid_to.strip()
            date_string = valid_to.replace(valid_to[-8:-5], '').rstrip()
            date_string = date_string.strip()
            valid_to = date_string

            datetime_object = datetime.strptime(
                date_string, '%a %b %d %H:%M:%S %Y').strftime("%Y-%m-%d %H:%M:%S")
            datetime_object = datetime.fromisoformat(datetime_object)
            currentDateTime = datetime.utcnow()
            expiringIn = str(datetime_object - currentDateTime)
            if expiringIn > '0':
                isExpired = False
            else:
                isExpired = True
    for sno in trustCert_detail:
        sno = sno.strip()
        if sno.startswith('Serial Number'):
            serialNumber = sno.strip('Serial Number:')
            serialNumber = serialNumber.strip()
    serviceName = cert_trust.split('/')[0]
    trust_cert_dict = {
        "name": cert_trust,
        "trackingId": serviceName + "_trust",
        "serialNumber" : serialNumber,
        "service": serviceName,
        "type": "TrustCertificate",
        "validFrom": valid_from,
        "validTo": valid_to,
        "expiringIn": expiringIn,
        "hasExpired": isExpired,
    }
    return trust_cert_dict


def fetch_trust_certs(remote_conn, trust_cert_list, trust_cert_dict, error_cert_dict):
    remote_conn.send('show cert list trust\n')
    time.sleep(10)
    trustCert = remote_conn.recv(50000)
    trustCert = trustCert.decode()
    trustCert = trustCert.split("\r\n")
    for eachTrustCert in trustCert:
        if ".pem" in eachTrustCert:
            trust_cert_name = eachTrustCert.split(":")[0]
            trust_cert_details = cert_trust_details(trust_cert_name, remote_conn)
            if trust_cert_details['validTo'] == None: 
                error_cert_dict.append(trust_cert_details)
                print("None Detected, Waiting 5 Seconds")
                time.sleep(5)
                remote_conn.send('\n')
            else:
                trust_cert_list.append(trust_cert_details)
    #print("Trust List  : ", trust_cert_list)
    return trust_cert_list, error_cert_dict


def formatFont(filename):
    ft = Font(color="00FF0000")
    wrbk = load_workbook(filename)
    for sh in wrbk.worksheets:
        rows = sh.max_row
        for i in range(2, rows):
            c = (sh.cell(row=i, column=9))
            if c.value == True:
                c.font = ft
        wrbk.save(filename)


def formatWidth(file):
    workbook = load_workbook(file)
    sheetsList = workbook.sheetnames

    for key in sheetsList:
        ws = workbook[key]
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
            workbook.save(file)

def getData(filename):
    for row in range(len(df)):
        hostname = df.ip[row]
        userName = df.userName[row]
        password = df.password[row]
        remote_conn_pre = paramiko.SSHClient()
        remote_conn_pre.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        remote_conn_pre.connect(hostname, username=userName, password=password, look_for_keys=False, allow_agent=False)
        remote_conn = remote_conn_pre.invoke_shell()
        remote_conn.send('\n')
        time.sleep(20)

        identity_cert_list = []
        trust_cert_list = []
        identity_cert_dict = {}
        trust_cert_dict = {}
        error_cert_dict = []

        identity_certificates = fetch_identity_certs(remote_conn, identity_cert_list, identity_cert_dict)
        # identity_certificates = [{'name': 'tomcat/tomcat.pem', 'trackingId': 'tomcat_id', 'fkService': 'tomcat', 'tktype': 'IdentityCertificate', 'validFrom': 'Fri Feb 04 15:31:39 PST 2022', 'validTo': 'Wed Feb 03 15:31:38  2027', 'expiringIn': '1776 days, 11:00:35.900868', 'hasExpired': True}, {'name': 'tomcat-ECDSA/tomcat-ECDSA.pem', 'trackingId': 'tomcat-ECDSA_id', 'fkService': 'tomcat-ECDSA', 'tktype': 'IdentityCertificate', 'validFrom': 'Fri Feb 04 15:31:41 PST 2022', 'validTo': 'Wed Feb 03 15:31:40  2027', 'expiringIn': '1776 days, 11:00:22.884183', 'hasExpired': False}, {'name': 'ipsec/ipsec.pem', 'trackingId': 'ipsec_id', 'fkService': 'ipsec', 'tktype': 'IdentityCertificate', 'validFrom': 'Fri Feb 04 15:31:34 PST 2022', 'validTo': 'Wed Feb 03 15:31:33  2027', 'expiringIn': '1776 days, 11:00:00.859514', 'hasExpired': True}, {'name': 'ITLRecovery/ITLRecovery.pem', 'trackingId': 'ITLRecovery_id', 'fkService': 'ITLRecovery', 'tktype': 'IdentityCertificate', 'validFrom': 'Fri Feb 04 15:31:45 PST 2022', 'validTo': 'Wed Feb 03 15:31:44  2027', 'expiringIn': '1776 days, 10:59:56.831187', 'hasExpired': False}, { 'name': 'CallManager-ECDSA/CallManager-ECDSA.pem', 'trackingId': 'CallManager-ECDSA_id', 'fkService': 'CallManager-ECDSA', 'tktype': 'IdentityCertificate', 'validFrom': 'Thu Feb 24 12:13:12 PST 2022', 'validTo': 'Tue Feb 23 12:13:11  2027', 'expiringIn': '1796 days, 7:41:08.814247', 'hasExpired': False}, {'name': 'CallManager/CallManager.pem', 'trackingId': 'CallManager_id', 'fkService': 'CallManager', 'tktype': 'IdentityCertificate', 'validFrom': 'Fri Feb 04 15:23:40 PST 2022', 'validTo': 'Wed Feb 03 15:23:39  2027', 'expiringIn': '1776 days, 10:51:21.792836', 'hasExpired': False}, {'name': 'CAPF/CAPF.pem', 'trackingId': 'CAPF_id', 'fkService': 'CAPF', 'tktype': 'IdentityCertificate', 'validFrom': 'Fri Feb 04 15:23:48 PST 2022', 'validTo': 'Wed Feb 03 15:23:47  2027', 'expiringIn': '1776 days, 10:51:14.770059', 'hasExpired': False}, {'name': 'TVS/TVS.pem', 'trackingId': 'TVS_id', 'fkService': 'TVS', 'tktype': 'IdentityCertificate', 'validFrom': 'Fri Feb 04 15:23:52 PST 2022', 'validTo': 'Wed Feb 03 15:23:51  2027', 'expiringIn': '1776 days, 10:51:03.751281', 'hasExpired': False}]
        # trust_certificates = [{'name': 'tomcat-trust/USERTrust_RSA_Certification_Authority_01fd6d30fca3ca51a81bbc640e35032d.pem', 'trackingId': 'tomcat-trust_trust', 'fkService': 'tomcat-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Sun Jan 31 16:00:00 PST 2010', 'validTo': 'Mon Jan 18 15:59:59  2038', 'expiringIn': '5778 days, 11:26:46.715123', 'hasExpired': True}, {'name': 'tomcat-trust/InCommon_RSA_Server_CA_4720d0fa85461a7e17a1640291846374.pem', 'trackingId': 'tomcat-trust_trust', 'fkService': 'tomcat-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Sun Oct 05 17:00:00 PDT 2014', 'validTo': 'Sat Oct 05 16:59:59  2024', 'expiringIn': '925 days, 12:26:31.689357', 'hasExpired': False}, {'name': 'tomcat-trust/cucm14-pub.collabiz.ext_791a41de2f3d66a9a4a560204d2feb5a.pem', 'trackingId': 'tomcat-trust_trust', 'fkService': 'tomcat-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri Feb 04 15:31:39 PST 2022', 'validTo': 'Wed Feb 03 15:31:38  2027', 'expiringIn': '1776 days, 10:57:55.673116', 'hasExpired': False}, {'name': 'tomcat-trust/cucm14-pub-EC.collabiz.ext_754b30f0e53d897fd1eddc84a85cf200.pem', 'trackingId': 'tomcat-trust_trust', 'fkService': 'tomcat-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri Feb 04 15:31:41 PST 2022', 'validTo': 'Wed Feb 03 15:31:40  2027', 'expiringIn': '1776 days, 10:57:42.643371', 'hasExpired': False}, {'name': 'CallManager-trust/CAPF-36205338.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri Feb 04 15:23:48 PST 2022', 'validTo': 'Wed Feb 03 15:23:47  2027', 'expiringIn': '1776 days, 10:49:34.613584', 'hasExpired': False}, {'name': 'CallManager-trust/Cisco_Root_CA_2099_019a335878ce16c1c1.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Tue Aug 09 13:58:28 PDT 2016', 'validTo': 'Sun Aug 09 13:58:28  2099', 'expiringIn': '28261 days, 9:24:00.592497', 'hasExpired': False}, {'name': 'CallManager-trust/Cisco_Manufacturing_CA_SHA2_02.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Mon Nov 12 05:50:58 PST 2012', 'validTo': 'Thu Nov 12 05:00:17  2037', 'expiringIn': '5711 days, 0:25:34.574499', 'hasExpired': True}, {'name': 'CallManager-trust/High_Assurance_SUDI_CA_0a6475524cd8617c62.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Thu Aug 11 13:28:08 PDT 2016', 'validTo': 'Sun Aug 09 13:58:27  2099', 'expiringIn': '28261 days, 9:23:29.560197', 'hasExpired': False}, {'name': 'CallManager-trust/Cisco_Basic_Assurance_Root_CA_2099_01a65af15ee994ebe1.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri May 26 12:19:29 PDT 2017', 'validTo': 'Tue May 26 12:19:29  2099', 'expiringIn': '28186 days, 7:44:16.540520', 'hasExpired': False}, {'name': 'CallManager-trust/ACT2_SUDI_CA_61096e7d00000000000c.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Thu Jun 30 10:56:57 PDT 2011', 'validTo': 'Mon May 14 13:25:42  2029', 'expiringIn': '2607 days, 8:50:14.533279', 'hasExpired': False}, {'name': 'CallManager-trust/Cisco_Root_CA_2048_5ff87b282b54dc8d42a315b568c9adff.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri May 14 13:17:12 PDT 2004', 'validTo': 'Mon May 14 13:25:42  2029', 'expiringIn': '2607 days, 8:49:59.518597', 'hasExpired': False}, {'name': 'CallManager-trust/Cisco_Manufacturing_CA_6a6967b3000000000003.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri Jun 10 15:16:01 PDT 2005', 'validTo': 'Mon May 14 13:25:42  2029', 'expiringIn': '2607 days, 8:49:44.488234', 'hasExpired': False}, {     'name': 'CallManager-trust/Cisco_Root_CA_M2_01.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Mon Nov 12 05:00:18 PST 2012', 'validTo': 'Thu Nov 12 05:00:18  2037', 'expiringIn': '5711 days, 0:24:05.468519', 'hasExpired': False}, {'name': 'CallManager-trust/Cisco_Manufacturing_CA_III_04302a0b364ce2da93.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Wed Jul 05 12:28:06 PDT 2017', 'validTo': 'Tue May 26 12:19:28  2099', 'expiringIn': '28186 days, 7:43:00.451050', 'hasExpired': False}, {'name': 'CallManager-trust/CCMPUB-EC_6fa80c1b26ef5c540529502ea9ade80b.pem', 'trackingId': 'CallManager-trust_trust', 'fkService': 'CallManager-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Sat Oct 19 23:45:34 PDT 2019', 'validTo': 'Thu Oct 17 23:45:33  2024', 'expiringIn': '937 days, 19:08:50.443291', 'hasExpired': False}, {'name': 'CAPF-trust/CAPF-36205338.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri Feb 04 15:23:48 PST 2022', 'validTo': 'Wed Feb 03 15:23:47  2027', 'expiringIn': '1776 days, 10:46:49.423839', 'hasExpired': False}, {'name': 'CAPF-trust/Cisco_Root_CA_2099_019a335878ce16c1c1.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Tue Aug 09 13:58:28 PDT 2016', 'validTo': 'Sun Aug 09 13:58:28  2099', 'expiringIn': '28261 days, 9:21:15.405614', 'hasExpired': False}, {'name': 'CAPF-trust/Cisco_Manufacturing_CA_SHA2_02.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Mon Nov 12 05:50:58 PST 2012', 'validTo': 'Thu Nov 12 05:00:17  2037', 'expiringIn': '5711 days, 0:22:49.385992', 'hasExpired': False}, {'name': 'CAPF-trust/High_Assurance_SUDI_CA_0a6475524cd8617c62.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Thu Aug 11 13:28:08 PDT 2016', 'validTo': 'Sun Aug 09 13:58:27  2099', 'expiringIn': '28261 days, 9:20:44.379928', 'hasExpired': False}, {'name': 'CAPF-trust/Cisco_Basic_Assurance_Root_CA_2099_01a65af15ee994ebe1.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri May 26 12:19:29 PDT 2017', 'validTo': 'Tue May 26 12:19:29  2099', 'expiringIn': '28186 days, 7:41:31.361205', 'hasExpired': False}, {'name': 'CAPF-trust/ACT2_SUDI_CA_61096e7d00000000000c.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Thu Jun 30 10:56:57 PDT 2011', 'validTo': 'Mon May 14 13:25:42  2029', 'expiringIn': '2607 days, 8:47:29.343505', 'hasExpired': False}, {'name': 'CAPF-trust/Cisco_Root_CA_2048_5ff87b282b54dc8d42a315b568c9adff.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri May 14 13:17:12 PDT 2004', 'validTo': 'Mon May 14 13:25:42  2029', 'expiringIn': '2607 days, 8:47:14.327786', 'hasExpired': False}, {'name': 'CAPF-trust/Cisco_Manufacturing_CA_6a6967b3000000000003.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Fri Jun 10 15:16:01 PDT 2005', 'validTo': 'Mon May 14 13:25:42  2029', 'expiringIn': '2607 days, 8:46:59.316620', 'hasExpired': False}, {'name': 'CAPF-trust/Cisco_Root_CA_M2_01.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Mon Nov 12 05:00:18 PST 2012', 'validTo': 'Thu Nov 12 05:00:18  2037', 'expiringIn': '5711 days, 0:21:20.286067', 'hasExpired': False}, {'name': 'CAPF-trust/Cisco_Manufacturing_CA_III_04302a0b364ce2da93.pem', 'trackingId': 'CAPF-trust_trust', 'fkService': 'CAPF-trust', 'tktype': 'TrustCertificate', 'validFrom': 'Wed Jul 05 12:28:06 PDT 2017', 'validTo': 'Tue May 26 12:19:28  2099', 'expiringIn': '28186 days, 7:40:15.277166', 'hasExpired': False}]
        # error_certificates = [{'name': 'tomcat-trust/IPTWTE1CA.bankofamerica.com.pem', 'trackingId': 'tomcat-trust_trust', 'service': 'tomcat-trust', 'type': 'TrustCertificate', 'validFrom': 'Fri Sep 22 03:15:26 EDT 2017', 'validTo': 'Sun Sep 22 03:15:26  2019', 'expiringIn': '-928days, 16:26:39.329653', 'hasExpired': True},{'name': 'tomcat-trust/Bank_of_America_Root_CA_-_G3.pem', 'trackingId': 'tomcat-trust_trust', 'service': 'tomcat-trust', 'type': 'TrustCertificate', 'validFrom': None, 'validTo': None, 'expiringIn': None, 'hasExpired': None},{'name': 'tomcat-trust/Infrastructure_Authority_Primary_-_G3.pem', 'trackingId': 'tomcat-trust_trust', 'service': 'tomcat-trust', 'type': 'TrustCertificate', 'validFrom': None, 'validTo': None, 'expiringIn': None, 'hasExpired': None}]
        identity_df = pd.DataFrame(identity_certificates)
        trust_certificates, error_certificates = fetch_trust_certs(remote_conn, trust_cert_list, trust_cert_dict, error_cert_dict)
        trust_df = pd.DataFrame(trust_certificates)
        error_df = pd.DataFrame(error_certificates)
    
        concatenated = pd.concat([identity_df, trust_df])
        
        if(len(error_df) > 0):
            with pd.ExcelWriter(error_path, engine='openpyxl', mode = 'a') as writer:
                error_df.to_excel(writer, sheet_name=hostname)

        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            concatenated.to_excel(writer, sheet_name=hostname)
        
    formatFont(output_path)
    formatWidth(output_path)

currentTime = str(datetime.now())
filename = input('Enter Input File Path: ')
df = pd.read_csv(filename + '.csv')
output_path = filename + '_output_'+currentTime + '.xlsx'   
output_writer = pd.ExcelWriter(output_path)
error_path = filename+"_error_"+currentTime+".xlsx"
error_writer = pd.ExcelWriter(error_path)
output_writer.save()
error_writer.save()

getData(filename)
